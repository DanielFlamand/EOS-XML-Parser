import xml.etree.ElementTree as ET
import openpyxl
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime

# ============================================================
#  INSTELLINGEN — pas dit eenmalig aan
# ============================================================
OUTPUT_MAP = r"C:\Users\dflamand\Ultimoo Groep\Proces & Innovatie - IT support en development - IT support en development\PowerQuery\EOS XML aanlevering\EOS XML Output Excel"
# ============================================================

def get_text(el, tag, default=''):
    if el is None:
        return default
    child = el.find(tag)
    return child.text.strip() if child is not None and child.text else default

def get_main_or_first(container, tag):
    items = container.findall(tag) if container is not None else []
    main = next((i for i in items if get_text(i, 'MAIN') == 'Y'), None)
    return main if main is not None else (items[0] if items else None)

def extract_debtor(debtor):
    addr     = get_main_or_first(debtor.find('ADDRESSLIST'), 'ADDRESS_TYPE')
    phone_el = get_main_or_first(debtor.find('PHONELIST'),   'PHONE_TYPE')
    email_el = get_main_or_first(debtor.find('EMAILLIST'),   'EMAIL_TYPE')
    return {
        'CASE_DEBTOR_ID':   get_text(debtor,   'CASE_DEBTOR_ID'),
        'DEBTOR_ROLE_CODE': get_text(debtor,   'DEBTOR_ROLE_CODE'),
        'DEBTOR_NAME1':     get_text(debtor,   'DEBTOR_NAME1'),
        'DEBTOR_NAME2':     get_text(debtor,   'DEBTOR_NAME2'),
        'GENDER':           get_text(debtor,   'GENDER'),
        'LEGAL_FORM':       get_text(debtor, 'LEGAL_FORM'),
        'BIRTH_DATE':       get_text(debtor, 'BIRTH_DATE'),
        'CONTACT_PERSON':   get_text(debtor, 'CONTACT_PERSON'),
        'ADDRESS_LINE1':    get_text(addr,     'LINE1'),
        'ZIPCODE':          get_text(addr,     'ZIPCODE'),
        'CITY':             get_text(addr,     'CITY'),
        'COUNTRY_CODE':     get_text(addr,     'COUNTRY_CODE'),
        'PHONE_NUMBER':     get_text(phone_el, 'PHONE_NUMBER'),
        'EMAIL':            get_text(email_el, 'EMAIL'),
    }

DEBTOR_FIELDS = [
    'CASE_DEBTOR_ID','DEBTOR_ROLE_CODE','DEBTOR_NAME1','DEBTOR_NAME2',
    'GENDER','LEGAL_FORM','BIRTH_DATE','CONTACT_PERSON','ADDRESS_LINE1','ZIPCODE','CITY','COUNTRY_CODE',
    'PHONE_NUMBER','EMAIL',
]

CASE_FIELDS = [
    'CASE_ID','PARTNER_CASE_ID','CLIENT_CASE_ID','BUSINESS_LINE',
    'TOTAL_BALANCE','LOCAL_CURRENCY','CASE_PRODUCT','CASE_PROCESS','CASE_CREATION_DATE',
    'ADDITIONAL_INFORMATION',
]

FIN_FIELDS = [
    'CASE_FINANCIAL_ID','FIN_STATUS','DOCUMENT_NUMBER','AMOUNT_EUR',
    'AMOUNT_DATE','DUE_DATE','FINANCIAL_CODE','DOCUMENT_CODE',
]

CLIENT_FIELDS = [
    'PCT_CLIENT_NAME','PCT_ADDITIONAL_INFORMATION',
]

SECTION_STYLES = {
    'case':   ('2E75B6', 'D6E4F0'),
    'debtor': ('375623', 'D5E8D4'),
    'fin':    ('7030A0', 'E1D5E7'),
    'client': ('C55A11', 'FCE4D6'),
}

COL_WIDTHS = {
    'CASE_ID':14,'PARTNER_CASE_ID':26,'CLIENT_CASE_ID':14,'BUSINESS_LINE':12,
    'TOTAL_BALANCE':14,'LOCAL_CURRENCY':10,'CASE_PRODUCT':12,'CASE_PROCESS':10,
    'CASE_CREATION_DATE':18,'ADDITIONAL_INFORMATION':40,'CASE_DEBTOR_ID':14,'DEBTOR_ROLE_CODE':10,
    'DEBTOR_NAME1':18,'DEBTOR_NAME2':16,'GENDER':8,'LEGAL_FORM':16,'BIRTH_DATE':16,'CONTACT_PERSON':20,
    'ADDRESS_LINE1':26,'ZIPCODE':10,'CITY':16,'COUNTRY_CODE':10,
    'PHONE_NUMBER':18,'EMAIL':26,'CASE_FINANCIAL_ID':16,'FIN_STATUS':10,
    'DOCUMENT_NUMBER':20,'AMOUNT_EUR':12,'AMOUNT_DATE':18,'DUE_DATE':18,
    'FINANCIAL_CODE':14,'DOCUMENT_CODE':12,
    'PCT_CLIENT_NAME':30,'PCT_ADDITIONAL_INFORMATION':50,
}

def maak_excel(xml_pad, output_map):
    tree = ET.parse(xml_pad)
    root = tree.getroot()

    max_debtors = max(
        (len(c.findall('.//DEBTORLIST/DEBTOR_TYPE')) for c in root.findall('.//CASE_TYPE')),
        default=1
    )
    max_debtors = max(max_debtors, 1)

    rows = []
    for case in root.findall('.//CASE_TYPE'):
        case_data = {
            'CASE_ID':            get_text(case, 'CASE_ID'),
            'PARTNER_CASE_ID':    get_text(case, 'PARTNER_CASE_ID'),
            'CLIENT_CASE_ID':     get_text(case, 'CLIENT_CASE_ID'),
            'BUSINESS_LINE':      get_text(case, 'BUSINESS_LINE'),
            'TOTAL_BALANCE':      get_text(case, 'TOTAL_BALANCE'),
            'LOCAL_CURRENCY':     get_text(case, 'LOCAL_CURRENCY'),
            'CASE_PRODUCT':       get_text(case, 'CASE_PRODUCT'),
            'CASE_PROCESS':       get_text(case, 'CASE_PROCESS'),
            'CASE_CREATION_DATE':     get_text(case, 'CREATION_DATE'),
            'ADDITIONAL_INFORMATION': get_text(case, 'ADDITIONAL_INFORMATION'),
        }

        debtors = case.findall('.//DEBTORLIST/DEBTOR_TYPE')
        debtor_data = [extract_debtor(d) for d in debtors]
        while len(debtor_data) < max_debtors:
            debtor_data.append({f: '' for f in DEBTOR_FIELDS})

        pct = case.find('PARTNERCLIENTTYPE')
        client_data = {
            'PCT_CLIENT_NAME':            get_text(pct, 'CLIENT_NAME'),
            'PCT_ADDITIONAL_INFORMATION': get_text(pct, 'ADDITIONAL_INFORMATION'),
        }

        fins = case.findall('.//CASEFINLIST/CASE_FIN_TYPE') or [None]
        for fin in fins:
            rows.append((case_data, debtor_data, {
                'CASE_FINANCIAL_ID': get_text(fin, 'CASE_FINANCIAL_ID'),
                'FIN_STATUS':        get_text(fin, 'STATUS'),
                'DOCUMENT_NUMBER':   get_text(fin, 'DOCUMENT_NUMBER'),
                'AMOUNT_EUR':        get_text(fin, 'AMOUNT_EUR'),
                'AMOUNT_DATE':       get_text(fin, 'AMOUNT_DATE'),
                'DUE_DATE':          get_text(fin, 'DUE_DATE'),
                'FINANCIAL_CODE':    get_text(fin, 'FINANCIAL_CODE'),
                'DOCUMENT_CODE':     get_text(fin, 'DOCUMENT_CODE'),
            }, client_data))

    # Kolommap opbouwen
    col_map = []
    for f in CASE_FIELDS:
        col_map.append(('Case', f, 'case'))
    for f in FIN_FIELDS:
        col_map.append(('Financiën', f, 'fin'))
    for f in CLIENT_FIELDS:
        col_map.append(('Crediteur', f, 'client'))
    for i in range(max_debtors):
        label = f'Debiteur {i+1}' if max_debtors > 1 else 'Debiteur'
        for f in DEBTOR_FIELDS:
            col_map.append((label, f, 'debtor'))

    # Excel opbouwen
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import"
    thin   = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hdr(row, col, value, section):
        hc, _ = SECTION_STYLES[section]
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        c.fill      = PatternFill('solid', start_color=hc)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = border

    def dat(row, col, value, section, alt):
        _, dc = SECTION_STYLES[section]
        c = ws.cell(row=row, column=col, value=value)
        c.font      = Font(name='Arial', size=10)
        c.fill      = PatternFill('solid', start_color=dc if alt else 'FFFFFF')
        c.alignment = Alignment(vertical='center', wrap_text=(section == 'client'))
        c.border    = border

    # Rij 1: sectielabels (gemerged)
    prev_label, merge_start = None, 1
    sections_order = [s for _, _, s in col_map]
    for ci, (label, _, section) in enumerate(col_map, 1):
        if label != prev_label:
            if prev_label is not None:
                if merge_start < ci - 1:
                    ws.merge_cells(start_row=1, start_column=merge_start,
                                   end_row=1, end_column=ci - 1)
                hdr(1, merge_start, prev_label, sections_order[merge_start - 1])
            merge_start, prev_label = ci, label
    if prev_label:
        last = len(col_map)
        if merge_start < last:
            ws.merge_cells(start_row=1, start_column=merge_start,
                           end_row=1, end_column=last)
        hdr(1, merge_start, prev_label, sections_order[merge_start - 1])
    ws.row_dimensions[1].height = 22

    # Rij 2: veldnamen
    for ci, (_, field, section) in enumerate(col_map, 1):
        hdr(2, ci, field, section)
    ws.row_dimensions[2].height = 28

    # Data
    for ri, (cd, dds, fd, cld) in enumerate(rows, 3):
        alt = (ri % 2 == 0)
        col = 1
        for f in CASE_FIELDS:
            dat(ri, col, cd[f], 'case', alt); col += 1
        for f in FIN_FIELDS:
            dat(ri, col, fd[f], 'fin', alt); col += 1
        for f in CLIENT_FIELDS:
            dat(ri, col, cld[f], 'client', alt); col += 1
        for dd in dds:
            for f in DEBTOR_FIELDS:
                dat(ri, col, dd.get(f, ''), 'debtor', alt); col += 1

    # Kolombreedte
    for ci, (_, field, _) in enumerate(col_map, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(field, 14)

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = f"A2:{get_column_letter(len(col_map))}2"

    timestamp    = datetime.now().strftime("%Y%m%d_%H%M%S")
    bestandsnaam = f"XML_Import_{timestamp}.xlsx"
    output_pad   = Path(output_map) / bestandsnaam
    Path(output_map).mkdir(parents=True, exist_ok=True)
    wb.save(output_pad)
    return output_pad, len(rows), max_debtors

def main():
    import sys
    from pathlib import Path

    INVOER_BESTAND = r"C:\Temp\eos_input.txt"

    # Bepaal het XML pad:
    # 1. Eerst kijken of er een invoerbestand is (vanuit PAD)
    # 2. Dan kijken of er een argument is meegegeven
    # 3. Anders popup tonen (lokale PC)

    xml_pad = None

    if Path(INVOER_BESTAND).exists():
        # Vanuit PAD via tekstbestand
        with open(INVOER_BESTAND, 'r', encoding='utf-8') as f:
            xml_pad = f.read().strip()
        # Invoerbestand verwijderen na uitlezen
        Path(INVOER_BESTAND).unlink()

    elif len(sys.argv) > 1:
        # Via commandoregel argument
        xml_pad = sys.argv[1].strip()

    if xml_pad:
        # PAD modus — geen popup, output naar console
        if not Path(xml_pad).exists():
            print(f"FOUT: Bestand niet gevonden: {xml_pad}")
            sys.exit(1)
        try:
            output_pad, aantal_rijen, aantal_debiteuren = maak_excel(xml_pad, OUTPUT_MAP)
            print(f"SUCCES")
            print(f"OUTPUT_PAD:{output_pad}")
            print(f"RIJEN:{aantal_rijen}")
            print(f"DEBITEUREN:{aantal_debiteuren}")
        except Exception as e:
            print(f"FOUT:{e}")
            sys.exit(1)
    else:
        # Lokale PC modus — popup tonen
        root = tk.Tk()
        root.withdraw()
        xml_pad = filedialog.askopenfilename(
            title="Selecteer het XML-bestand",
            filetypes=[("XML bestanden", "*.xml"), ("Alle bestanden", "*.*")]
        )
        if not xml_pad:
            messagebox.showinfo("Geannuleerd", "Geen bestand geselecteerd.")
            return
        try:
            output_pad, aantal_rijen, aantal_debiteuren = maak_excel(xml_pad, OUTPUT_MAP)
            messagebox.showinfo(
                "Klaar!",
                f"Excel aangemaakt:\n{output_pad}\n\n"
                f"Rijen: {aantal_rijen}\n"
                f"Debiteur-groepen per rij: {aantal_debiteuren}"
            )
        except Exception as e:
            messagebox.showerror("Fout", f"Er ging iets mis:\n{e}")

if __name__ == "__main__":
    main()
